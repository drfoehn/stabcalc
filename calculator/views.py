import base64
from tempfile import NamedTemporaryFile
import pandas
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from django.core.exceptions import ValidationError
from django.core.files.storage import FileSystemStorage
from django.core.mail import EmailMessage, send_mail, BadHeaderError
from django.forms import modelformset_factory  # is grey but still needed for the result_add_view
from datetime import datetime
from wsgiref.util import FileWrapper
import json
from django.template.loader import render_to_string
from django.utils import timezone
from django.urls import reverse_lazy

from django.utils.html import format_html
from django.utils.translation import gettext_lazy as _
from statsmodels.stats.power import TTestIndPower
from django.views.generic import (
    ListView,
    DetailView,
    CreateView,
    UpdateView,
    TemplateView, FormView,
)

from .filters import SettingFilter, ResultFilter
from .forms import *
from django.shortcuts import redirect, render, get_object_or_404
from .models import *
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
from matplotlib.figure import Figure
import plotly.express as px
import plotly.graph_objects as go
import seaborn as sns
import seaborn.objects as so
import sklearn
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import PolynomialFeatures, SplineTransformer
import statsmodels.api as sm
import math
from patsy.highlevel import dmatrices
from django.http import HttpResponse, HttpResponseForbidden, HttpResponseRedirect
from io import BytesIO
from openpyxl import Workbook, load_workbook  # Documentation at https://openpyxl.readthedocs.io/en/stable/tutorial.html
from openpyxl.cell import cell
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

class ResultsAdminView(DetailView):
    model = Result
    template_name = 'calculator/results_admin.html'

class ResultsView(DetailView):
    template_name = "calculator/results.html"
    model = Setting
    context_object_name = "setting"

    def get_context_data(self, **kwargs):
        global single_results_value, single_results_duration
        context = super().get_context_data(**kwargs)
        subjects = Subject.objects.filter(settings__in=[self.object])
        durations = Duration.objects.filter(settings__in=[self.object])
        durations_val = durations.values()
        durations_data = pd.DataFrame(durations_val)
        results = Result.objects.filter(setting=self.object)
        results_val = results.values()
        results_data = pd.DataFrame(results_val)
        parameter = ParameterUser.objects.filter(setting=self.object).values('parameter__name', 'parameter__unit')
        setting = self.object
        context['subjects_n'] = Subject.objects.filter(settings__in=[self.object]).count()
        context['durations_n'] = Duration.objects.filter(settings__in=[self.object]).count()
        context['results_n'] = Result.objects.filter(setting=self.object).count()

        # ---------------------------Get deviation data for each subject setting and time individually in a dict
        deviation_dict: dict[int, dict[int, int]] = {}
        for subject in subjects:
            if not subject.id in deviation_dict:
                deviation_dict[subject.id] = {}
                for duration in self.object.durations.all():
                    if setting == self.object:
                        deviation_dict[subject.id][duration.seconds] = subject.deviation(duration, setting)

        # ----- trasform deviation-Dictionary into 2D-Array for SKLearn

        duration_list=[]
        result_list = []
        for key, value in deviation_dict.items():
            for duration, result in value.items():
                if duration == None or result == None:
                    pass
                else:
                    duration_list.append(duration/3600)
                    result_list.append(result)


        # result_arr = np.array(result_list).reshape(-1,1)
        # duration_arr = np.array(duration_list).reshape(-1,1)



        deviation_array = pd.DataFrame(deviation_dict)
        deviation_array.index.name = "duration"

        # https://www.delftstack.com/howto/python-pandas/how-to-iterate-through-rows-of-a-dataframe-in-pandas/



        # ------------Relative Deviation from Baseline
        y_rel = []
        x1_rel = []
        for (duration, results) in deviation_array.iterrows():
            for result in results:
                if not math.isnan(result):  # pandas converts None-values to "nan" - this function excludes those values
                    y_rel.append(result)
                    x1_rel.append(duration)

        # ---convert x (seconds) into hours
        x1_rel_hours = []
        for x in x1_rel:
            if x == 0:
                x_hour = 0
            else:
                x_hour = x / 3600
            x1_rel_hours.append(x_hour)

        #------preparing data for regression analysis and plotting
        stor_dur = x1_rel_hours
        stor_dur_arr = np.array(stor_dur).reshape(-1,1)
        stor_dev = y_rel
        stor_dev_arr = np.array(stor_dev).reshape(-1, 1)
        zipped_lin = list(zip(stor_dur, stor_dev))
        df_lin = pd.DataFrame(zipped_lin, columns=['Duration', 'Deviation'])

        stor_dur_square = []
        for dur in stor_dur:
            dur_square = dur**2
            stor_dur_square.append(dur_square)

        zipped_poly = list(zip(stor_dur_square, stor_dev))
        df_poly = pd.DataFrame(zipped_poly, columns=['Duration', 'Deviation'])


        # ---------Linear Regression with SKLearn
        lin_regr = LinearRegression(fit_intercept=False)
        lin_regr.fit(stor_dur_arr, stor_dev_arr)
        prediction = lin_regr.predict(np.sort(stor_dur_arr, axis=0))
        intercept_lin = lin_regr.intercept_
        r2_linregr = lin_regr.score(stor_dur_arr, stor_dev_arr)
        coeff_lin_1 = round(lin_regr.coef_[0][0], 2)
        context['r2_linregr'] = r2_linregr
        context['eq_linregr'] = "PD% = " + str(coeff_lin_1) + " * storage duration"


        #-------------Linear regression Graph
        sns.set_style('whitegrid')
        lin_plot = sns.lmplot(x='Duration', y='Deviation', data=df_lin)
        lin_plot_file = BytesIO()
        lin_plot.figure.savefig(lin_plot_file, format='png')
        b64 = base64.b64encode(lin_plot_file.getvalue()).decode()
        context['chart_lin'] = b64


   


        # sns.regplot().add(so.Line(), so.PolyFit())
        # lin_plot_file = BytesIO()
        # lin_plot.figure.savefig(lin_plot_file, format='png')
        # b64 = base64.b64encode(lin_plot_file.getvalue()).decode()
        # fig, ax = plt.subplots()
        # sns.boxplot(x='Duration', y='Deviation', data=df_lin, ax=ax)
        # sns.regplot(x='Duration', y='Deviation', data=df_lin, ax=ax, scatter=False)
        # ax.set_title('Linear Regression')
        # ax.set_xlabel("Storage Duration")
        # ax.set_ylabel("Deviation (PD%)")
        # flike = BytesIO()
        # fig.savefig(flike)
        # b64 = base64.b64encode(flike.getvalue()).decode()

        # context['chart_lin'] = b64


        # ---------Polynomial Regression 2nd° with SKLearn
        poly = PolynomialFeatures(degree=2, include_bias=False) # include_bias=False means that we deliberately want the y intercept (ß0) to be equal to 0
        stor_dur_poly = poly.fit_transform(stor_dur_arr) # create x2 values from our x values
        poly_regr = LinearRegression(fit_intercept=False)
        poly_regr.fit(stor_dur_poly, stor_dev_arr)
        prediction_poly = poly_regr.predict(np.sort(stor_dur_poly, axis=0))
        intercept_poly = poly_regr.intercept_
        r2_polyregr = poly_regr.score(stor_dur_poly, stor_dev_arr)
        coeff_poly_1 = round(poly_regr.coef_[0][0], 2)
        coeff_poly_2 = round(poly_regr.coef_[0][1], 2)
        context['r2_polyregr'] = r2_polyregr
        context['eq_polyregr'] = "PD% = " + str(coeff_poly_2) + " * storage duration^2 + " + str(coeff_poly_1) + "* storage duration"

        #-------------Polynomial 2nd° regression Graph
        sns.set_style('whitegrid')
        poly_plot = sns.lmplot(x='Duration', y='Deviation', data=df_lin, order=2)
        poly_plot_file = BytesIO()
        poly_plot.figure.savefig(poly_plot_file, format='png')
        b64_poly = base64.b64encode(poly_plot_file.getvalue()).decode()
        context['chart_poly'] = b64_poly


        # ----------------------Absolute values
        # ----------------------Merge data absolute results + duration


        merged_res_dur = pd.merge(
            results_data,
            durations_data,
            left_on="duration_id",
            right_on="id",
            how="inner",
        )


        merged_res_dur['hours'] = (merged_res_dur['seconds'] / 3600)
        context["results_data222"] = merged_res_dur.to_html

        # ----------------Numpy-Arrays
        context["results_array"] = np.array(merged_res_dur)

        # -------------------Variable absolute results
        y_abs = merged_res_dur["value"]  # dependent variable
        x1_abs = merged_res_dur["seconds"]  # independent variable


        ################################# REGRESSION ANALYSIS ######################################

        # https://365datascience.com/tutorials/python-tutorials/linear-regression/

        ## --------------------Explanation of statistics summary:
        # https://medium.com/swlh/interpreting-linear-regression-through-statsmodels-summary-4796d359035a
        # TODO: But why are there four different versions of Region when we only input one? Simply put, the formula expects continuous values in the form of numbers. By inputting region with data points as strings, the formula separates each string into categories and analyzes the category separately. Formatting your data ahead of time can help you organize and analyze this properly.

        # -----------------------LINEAR----------------------------------------

        vars = ['value', 'seconds']
        # print(merged_res_dur)
        merged_res_dur = merged_res_dur[vars]
        # print(merged_res_dur)
        merged_res_dur = merged_res_dur.dropna()
        # print(merged_res_dur)
        y=merged_res_dur['value']
        X=merged_res_dur['seconds']

        X = sm.add_constant(X)
        # y, X = dmatrices('value~seconds', data=merged_res_dur, return_type='dataframe')

        mod = sm.OLS(y, X)  # Describe model
        # mod = sm.OLS(formula="seconds ~ value", data=merged_res_dur)
        res = mod.fit()  # Fit model
        # print(res.summary())



        context["statistics_extended_abs_lin"] = res.summary()

        r_squared_lin = res.rsquared
        context["r_squared_lin"] = r_squared_lin
        r_squared_lin_adj = res.rsquared_adj
        context["r_squared_lin_adj"] = r_squared_lin_adj
        X = merged_res_dur.iloc[:, 1]  # seconds
        y = merged_res_dur.iloc[:, 0]  # value

        # Calculate Regression equation - linear
        eq1 = np.poly1d(np.polyfit(x1_rel_hours, y_rel, 1))

        context["eq_model_lin"] = str("y = " + str(round(eq1[1], 5)) + " * x + " + str(round(eq1[0], 5)))

        # prediction_lin= sm.ols("y ~ x", data=merged_res_dur).fit()
        # prediction_lin.predict(exog=new_values_dict)

        # --------------------------Calculate Regression equation - polynomial 2rd degree - https://www.statology.org/polynomial-regression-python/

        eq2 = np.poly1d(np.polyfit(x1_rel_hours, y_rel, 2))
        context["eq_model_poly2"] = str("y = " + str(round(eq2[2], 5))
                                        + " * x^2 + " + str(round(eq2[1], 5)) + " * x + " + str(round(eq2[0], 5)))
        # FIXME: Only the last two numbers get rounded on 5 digits ?!?!?

        # Calculate r_squared
        results = {}
        coeffs = np.polyfit(X, y, 2)
        p = np.poly1d(coeffs)

        # calculate r-squared
        yhat = p(X)
        ybar = np.sum(y) / len(y)
        ssreg = np.sum((yhat - ybar) ** 2)
        sstot = np.sum((y - ybar) ** 2)
        r_squared_poly2 = ssreg / sstot

        context["r_squared_poly2"] = r_squared_poly2

        # ---------------------------------Calculate Regression equation - polynomial 3rd degree

        eq3 = np.poly1d(np.polyfit(x1_rel_hours, y_rel, 3))
        context["eq_model_poly3"] = str("y = " + str(round(eq3[3], 5)) + " * x^3 + " + str(round(eq3[2], 5))
                                        + " * x^2 + " + str(round(eq3[1], 5)) + " * x + " + str(round(eq3[0], 5)))

        # Calculate r_squared
        results = {}
        coeffs = np.polyfit(X, y, 3)
        p = np.poly1d(coeffs)
        # calculate r-squared
        yhat = p(X)
        ybar = np.sum(y) / len(y)
        ssreg = np.sum((yhat - ybar) ** 2)
        sstot = np.sum((y - ybar) ** 2)
        r_squared_poly3 = ssreg / sstot

        context["r_squared_poly3"] = r_squared_poly3

        # TODO: Predicting values from regression: https://towardsdatascience.com/linear-regression-with-python-and-numpy-25d0e1dd220d
        # TODO:  ANOVA: https://www.statsmodels.org/dev/examples/notebooks/generated/interactions_anova.html

        # -----------------------LOG-LINEAR----------------------------------------

        def log_func(x):
            if not x:
                return 0
            else:
                x = np.log(x)
            return x

        merged_res_dur['seconds'] = merged_res_dur['seconds'].apply(log_func)
        y, X_log = dmatrices('value~seconds', data=merged_res_dur, return_type='dataframe')
        mod_log = sm.OLS(y, X_log)  # Describe model
        res_log = mod_log.fit()  # Fit model
        context["statistics_extended_log"] = res_log.summary()

        r_squared_log = res_log.rsquared
        context["r_squared_log"] = r_squared_log
        r_squared_log_adj = res_log.rsquared_adj
        context["r_squared_log_adj"] = r_squared_log_adj
        #

        # --------Log seconds and mean values for graph

        deviation_df = pd.DataFrame(deviation_array)
        deviation_df['mean_rows'] = deviation_df.mean(axis=1)
        deviation_df.reset_index(inplace=True)
        deviation_df['duration'] = deviation_df['duration'].apply(log_func)
        deviation_mean_df = deviation_df[['duration', 'mean_rows']].values

        dev_log_data = []
        for row in deviation_mean_df:
            dev_log_data.append([row[0], row[1]])

        context["log_data"] = dev_log_data

        # -------Calculate Regression equation - lin log
        # y = a + b*ln(x)
        # -- Reshape Dataframe into 1D-Array for calculations
        vars = ['seconds']
        X_log = X_log[vars]

        X_log_a = np.array(X_log)
        X_log_1D = X_log_a.ravel()
        y_a = np.array(y)
        y_1D = y_a.ravel()




        # --Calcualting equation
        eq_model_log = np.polyfit(X_log_1D, y_1D, 1)
        context["eq_model_log"] = str(
            "y = " + str(round(eq_model_log[0], 5)) + " * log(x) + " + str(round(eq_model_log[1], 5)))

        #############################Overall statistical tests for the datatable ###########################
        # Rainbow test for linearity (the null hypothesis is that the relationship is properly modelled as linear);
        # first number is an F-statistic and that the second is the p-value.

        rainbow_test = sm.stats.linear_rainbow(res)
        context["rainbow_f"] = rainbow_test[0]
        context["rainbow_p"] = rainbow_test[1]

        # print(rainbow_test)

        # Test assumed normal or exponential distribution using Lilliefors’ test.        #
        # Lilliefors’ test is a Kolmogorov - Smirnov test with estimated parameters.
        # first number: ksstatfloat Kolmogorov - Smirnov test  statistic with estimated mean and variance.
        # second number: pvaluefloat   If  the pvalue is lower  than some threshold, e.g. 0.05, then  we can  reject the   Null hypothesis  that  the  sample comes from a normal distribution.

        ks = sm.stats.diagnostic.kstest_normal(y, dist='norm', pvalmethod='table')
        context["ksstat"] = ks[0]
        ksstat_p = ks[1]
        context["ksstat_p"] = ks[1]
        # TODO. Provide explanation
        # kurtosistest only valid for n>=20

        # ------------------------Absolute

        x1_abs = sm.add_constant(x1_abs)  # add a row of ones as constant  https://365datascience.com/tutorials/python-tutorials/linear-regression/
        model = sm.OLS(y_abs, x1_abs)
        results_abs = model.fit()  # OLS = Ordinary Least Squares
        context["statistics_extended_abs_lin"] = results_abs.summary()

        # -------------------------Extract single parameters from summary - linear regression
        b0_abs_lin = results_abs.params[0]  # constant coefficient / Intercept
        b1_abs_lin = results_abs.params[1]  # seconds coefficient / Slope

        b0_r_abs_lin = round(b0_abs_lin, 5)
        context["intercept_abs_lin"] = b0_r_abs_lin
        b1_r_abs_lin = round(b1_abs_lin, 5)
        context["slope_abs_lin"] = b1_r_abs_lin
        context["reg_eq_abs_lin"] = (
                "y = " + str(b0_r_abs_lin) + " + x1 * " + str(b1_r_abs_lin)
        )

        # --------------------------relative

        x_rel = sm.add_constant(x1_rel)  # add a row of ones as constant
        model = sm.OLS(y_rel, x_rel)
        results_rel = model.fit()  # OLS = Ordinary Least Squares


        # # -------------------------Extract single parameters from summary - linear regression
        # b0_rel_lin = results_rel.params[0]  # constant coefficient / Intercept
        # b1_rel_lin = results_rel.params[1]  # seconds coefficient / Slope
        # b0_r_rel_lin = round(b0_rel_lin, 5)
        # context["intercept_rel_lin"] = b0_r_rel_lin
        # b1_r_rel_lin = round(b1_rel_lin, 5)
        # context["slope_rel_lin"] = b1_r_rel_lin
        # context["reg_eq_rel_lin"] = (
        #         "y = " + str(b0_r_rel_lin) + " + x1 * " + str(b1_r_rel_lin)
        # )

        ######################################INTERPRETATION ################################################

        # --------------------------Calculate best fitting model

        r_sq_list = [r2_linregr, r2_polyregr]
        best_fit = max(r_sq_list)
        context["best_fit"] = str(str(round(best_fit * 100, 2)) + " %")

        def best_fit_model() -> str:
            if r2_linregr == best_fit:
                return 'Linear Regression Model'
            # elif r2_polyregr == best_fit:
            else:
                return 'Polynomial Regression Model 2° degree'

        context["best_fit_model"] = best_fit_model()
        
        one_hour_lin = coeff_lin_1
        one_hour_poly = coeff_poly_1+coeff_poly_2

        context["interpretation_1"] = 'Under these conditions, a ' + str(
            parameter.values('parameter__name')[0]['parameter__name']) + ' magnitude increase/decrease of ' + str(
            round(one_hour_lin, 3)) + '% per hour is expected'

        context["interpretation_2"] = '1 hour of sample storage under the tested conditions causes the the ' + str(
            parameter.values('parameter__name')[0]['parameter__name']) + ' level to change by ' + str(round(one_hour_lin, 3)) + '%'

        # print(res.params[1])

        # -------------------------- Normal distribution

        def distrib() -> str:
            if ksstat_p <= 0.05:
                return 'The data IS NOT normally distributed. KS-p-value: ' + str(
                    round(ksstat_p, 4)) + ' (only valid if n > 20)'
            else:
                return 'The data IS normally distributed. KS-p-value: ' + str(
                    round(ksstat_p, 4)) + ' (only valid if n > 20)'

        context["interpretation_dist"] = distrib()

        ###################################  Power Analysis #########################################

        #  ---------------- power analysis - linear regression
        effect_lin = r2_linregr
        effect_poly = r2_polyregr
        effect_log = r_squared_log
        alpha = 0.05
        nobs = Subject.objects.filter(settings__in=[self.object]).count()

        # perform power analysis
        analysis = TTestIndPower()
        power_lin = analysis.solve_power(effect_lin, power=None, nobs1=nobs, ratio=1.0, alpha=alpha)
        power_poly = analysis.solve_power(effect_poly, power=None, nobs1=nobs, ratio=1.0, alpha=alpha)
        power_log = analysis.solve_power(effect_log, power=None, nobs1=nobs, ratio=1.0, alpha=alpha)

        context["power_lin"] = round(power_lin, 2)
        context["power_poly"] = round(power_poly, 2)
        context["power_log"] = round(power_log, 2)

        power_lin_est = []
        for lin_est in range(1, 10):
            lin_est = lin_est / 10
            # print(lin_est)
            nobs = analysis.solve_power(effect_lin, power=lin_est, nobs1=None, ratio=1.0, alpha=alpha)
            data_for_graph = (nobs, lin_est)
            power_lin_est.append(data_for_graph)
        context['power_lin_est'] = power_lin_est

        # power = TTestIndPower().solve_power(effect_size=effect_size,
        #                                     nobs1=nobs1,
        #                                     ratio=ratio,
        #                                     power=None,
        #                                     alpha=alpha,
        #                                     alternative='two-sided')

        # ---------------------Image for Power estimation
        # -----------------------Render with <img src="data:image/png,base64,{{ power|safe }}" alt="">
        # nobs = Subject.objects.filter(settings__in=[self.object]).count()
        # print(nobs)
        # panalysis = TTestIndPower()
        #
        # def powerFigure():
        #     panalysis.plot_power(
        #         dep_var="nobs",
        #         nobs=np.arange(5, nobs),
        #         effect_size=np.arange(0.5, 1.5, .2),
        #         alpha=0.05,
        #         ax=None,
        #         title='Power-analysis',
        #
        #     )
        #     plt.xlabel('Number of subjects', fontsize=20)
        #     plt.ylabel('Power', fontsize=20)
        #     plt.show()
        #
        # context['power'] = panalysis
        # context['panalysis'] = panalysis

        ############################### Calculate maximum permissible error (MPE)##########################

        cv_g = ParameterUser.objects.filter(setting=self.object).values('parameter__cv_g')[0]['parameter__cv_g']
        cv_i = ParameterUser.objects.filter(setting=self.object).values('parameter__cv_i')[0]['parameter__cv_i']
        cv_a = ParameterUser.objects.filter(setting=self.object).values('cv_a')[0]['cv_a']
        context['cv_g'] = cv_g
        context['cv_i'] = cv_i
        context['cv_a'] = cv_a
        rcv = 2 ** 0.5 * (1.96 * (cv_a ** 2 + cv_i ** 2) ** 0.5)
        context['rcv'] = round(rcv, 2)
        allow_dev = 0.5 * cv_i
        context['allow_dev'] = round(allow_dev, 2)
        accept_dev = 0.7 * allow_dev
        context['accept_dev'] = round(accept_dev, 2)
        allow_bias = 0.25 * math.sqrt(cv_i ** 2 + cv_g ** 2)
        context['allow_bias'] = round(allow_bias, 2)

        rcv_mpe_zero = ((rcv) / coeff_lin_1)  # forced through zero
        context["rcv_mpe_zero"] = rcv_mpe_zero

        rcv_mpe_poly_zero = ((-coeff_poly_1+math.sqrt(coeff_poly_1**2 + 4*coeff_poly_2*rcv))/(2*coeff_poly_2))
        context["rcv_mpe_poly_zero"] = rcv_mpe_poly_zero
        print(rcv_mpe_poly_zero)

        ###################################  Data import / export ####################################

        # ---------------------------Import
        # def upload_stability_data(request):
        #     if request.method == "POST":
        #         form = UploadExcelForm(request.POST, request.FILES)
        #         if form.is_valid():
        #             kappa = request.FILES['file']
        #             # Do work on kappa/excel file here with Pandas
        #             output = io.BytesIO()
        #             writer = pd.ExcelWriter(output, engine='xlsxwriter')
        #             kappa.to_excel(writer, index=False)
        #             writer.save()
        #             output.seek(0)
        #             response = HttpResponse(output,
        #                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        #         response['Content-Disposition'] = 'attachment; filename=%s.xlsx' % 'Download'
        #         return response
        #     else:
        #         form = UploadExcelForm()
        #
        #     return render(request, 'calculator/upload_form.html', {'form': form})

        # -----------------------------Export

        # https://stackoverflow.com/questions/35267585/django-pandas-to-http-response-download-file
        # https://djangoadventures.com/how-to-create-file-download-links-in-django/
        #

        MDATA = datetime.now().strftime('%Y-%m-%d')

        def data_export_xlsx(pk):
            data = merged_res_dur
            slug = parameter.values('name')[0]['name']

            df = pd.DataFrame(merged_res_dur)
            # Convertendo data para o formato correto
            # df['dataset__created'] = df['dataset__created'].apply(
            #     lambda x: x.strftime('%Y-%m-%d %H:%M'))
            # df['dataset__modified'] = df['dataset__modified'].apply(
            #     lambda x: x.strftime('%Y-%m-%d %H:%M'))

            filename = f'stability_data_{slug}_{MDATA}.xlsx'

            import tempfile
            with tempfile.NamedTemporaryFile(suffix='.xlsx') as tmp:
                df.to_excel(tmp.name, sheet_name=f'stability_data_{slug}')
                tmp.flush()
                tmp.seek(0)
                data = tmp.read()
            return (data, filename)

        # context["export_data_excel"] = data_export_xlsx(id(self))

        # --------------------MatPlotLib Scattergramm:
        # plt.scatter(x1_rel, y_rel)
        # plt.xlabel('Seconds', fontsize=20)
        # plt.ylabel('Results', fontsize=20)
        # plt.show()

        # ------------------------Return context

        context.update(
            {
                "subjects": subjects,
                "results": results,
                "durations": Duration.objects.all(),
                "r_square": results_abs.rsquared,
                "f_p_value": results_abs.f_pvalue,
                "f_p_value_perc": results_abs.f_pvalue * 100,
                "f_value": results_abs.fvalue,
                "f_p_log_value": res_log.f_pvalue,
                "f_p_log_value_perc": res_log.f_pvalue * 100,
                # Essentially, it asks, is this a useful variable? Does it help us explain the variability we have in this case?
            }
        )
        return context






def DownloadExcel(request, setting_pk):
    excelfile = BytesIO()
    workbook = Workbook()
    workbook.remove(workbook.active)
    owner = request.user
    setting = Setting.objects.get(pk=setting_pk, owner=owner)
    durations = setting.durations.all()
    results = setting.results.all()
    print(results)
    ws1 = workbook.create_sheet(title="Basic Information")
    ws1.sheet_view.showGridLines = False

    now = timezone.now().strftime('%d-%m-%Y - %H:%M')

    ws1['A1'] = "EFLM Stability Calculator"
    ws1['A1'].font = Font(size=25, color="e63946", bold=True)
    ws1['A2'] = "EFLM Working Group Preanalytical Phase (WG-PRE)"
    ws1['A2'].font = Font(size=20, color="e63946")
    ws1['A4'] = 'Downloaded by ' + request.user.user_name + ' on ' + now
    ws1['A4'].font = Font(italic=True)
    ws1['A6'] = "Stability study setting details"
    ws1['A6'].font = Font(size=20, bold=True, color="1d3557")
    title3 = Font(size=18, color='1d3557')
    for c in ws1['A7:B8']:
        c[0].font = title3
        c[1].font = title3
    ws1['A7'] = "Name"
    ws1['B7'] = str(setting.name)
    ws1['A8'] = "Parameter"
    ws1['B8'] = str(setting.parameter)
    ws1['A9'] = "Samples"
    ws1['B9'] = str(setting.sample)
    ws1['A10'] = "Source of samples"
    ws1['B10'] = str(setting.get_sample_type_display)
    ws1['A11'] = "Storage Condition"
    ws1['B11'] = str(setting.condition)
    ws1['A12'] = "Type of study design"
    ws1['B12'] = str(setting.get_design_type_display())
    ws1['A13'] = "Primary samples or aliquots"
    ws1['B13'] = str(setting.get_design_sample_display())
    ws1['A14'] = "Freeze thaw cycles (n)"
    ws1['B14'] = str(setting.freeze_thaw_cycles)
    ws1['A15'] = "Protocol"
    ws1['B15'] = str(setting.protocol)
    ws1['A16'] = "Comment"
    ws1['B16'] = str(setting.comment)

    bold = Font(bold=True)
    for c in ws1['A9:A16']:
        c[0].font = bold

    thin = Side(border_style="thin", color="000000")
    bottom_line = Border(bottom=thin)
    for c in ws1['A9:B16']:
        c[0].border = bottom_line
        c[1].border = bottom_line

    MIN_WIDTH = 15
    for i, column_cells in enumerate(ws1.columns, start=1):
        width = (
            length
            if (length := max(len(str(cell_value) if (cell_value := cell.value) is not None else "")
                              for cell in column_cells)) >= MIN_WIDTH
            else MIN_WIDTH
        )
        ws1.column_dimensions[get_column_letter(i)].width = width

    # -------------sheet 2 - data ------------------

    ws2 = workbook.create_sheet('Data')
    ws2.sheet_view.showGridLines = False
    # for subject in setting.subjects.all():
    #     ws2.append([str(subject)])
    #     for duration in setting.durations.all():
    #         ws2.append([str(duration)])
    #         for result in setting.durations.
    ws2["A1"] = 'Subject'
    ws2["B1"] = 'Storage duration'
    ws2["C1"] = 'Result'
    ws2["D1"] = 'Unit'

    for c in ws2['A1:D1'][0]:
        c.font = bold
        c.border = bottom_line

    for count, result in enumerate(results):
        ws2.cell(row=count + 2, column=1).value = str(result.subject)
        ws2.cell(row=count + 2, column=2).value = str(result.duration)
        ws2.cell(row=count + 2, column=3).value = result.value
        ws2.cell(row=count + 2, column=4).value = str(result.setting.parameter.parameter.unit)

        MIN_WIDTH = 10
        for i, column_cells in enumerate(ws2.columns, start=1):
            width = (
                length
                if (length := max(len(str(cell_value) if (cell_value := cell.value) is not None else "")
                                  for cell in column_cells)) >= MIN_WIDTH
                else MIN_WIDTH
            )
            ws2.column_dimensions[get_column_letter(i)].width = width

    # -------------sheet 3 - statistics ------------------

    ws3 = workbook.create_sheet('Statistics')
    ws3.sheet_view.showGridLines = False

    ws3['A1'] = 'Descriptive statistics'
    ws3['A1'].font = Font(size=15, bold=True)
    ws3['A2'] = 'Number of Subjects'
    ws3['B2'] = Subject.objects.filter(settings__in=[setting]).count()
    ws3['A3'] = 'Number of tested storage durations'
    ws3['B3'] = Duration.objects.filter(settings__in=[setting]).count()
    ws3['A4'] = 'Number of results '
    ws3['B4'] = Result.objects.filter(setting=setting).count()

    for count, duration in enumerate(setting.durations.all()):
        ws3.cell(row=6, column=count + 2).value = str(duration)
        ws3.cell(row=6, column=count + 2).font = bold

    ws3['A7'] = 'Average'
    ws3['A8'] = 'Standard deviation low'
    ws3['A9'] = 'Standard deviation high'
    ws3['A10'] = 'Coefficient of Variation (%)'
    ws3['A11'] = 'Percent difference from baseline (%)'

    for count, duration in enumerate(setting.durations.all()):
        ws3.cell(row=7, column=count + 2).value = setting.average_tot(duration=duration)
    for count, duration in enumerate(setting.durations.all()):
        ws3.cell(row=8, column=count + 2).value = setting.avg_tot_sd_l(duration=duration)
    for count, duration in enumerate(setting.durations.all()):
        ws3.cell(row=9, column=count + 2).value = setting.avg_tot_sd_h(duration=duration)
    for count, duration in enumerate(setting.durations.all()):
        ws3.cell(row=10, column=count + 2).value = setting.cv_tot(duration=duration)
    for count, duration in enumerate(setting.durations.all()):
        ws3.cell(row=11, column=count + 2).value = setting.deviation_tot(duration=duration)

    for c in ws3['A7:A11']:
        c[0].font = bold

    ws3['A13'] = "Regression Analysis"
    ws3['A13'].font = Font(size=15, bold=True)
    ws3['A15'] = 'Add regression formula, R2, etc???'

    ws3['A23'] = 'Maximal Permissible Error'
    cv_g = ParameterUser.objects.filter(setting=setting).values('parameter__cv_g')[0]['parameter__cv_g']
    cv_i = ParameterUser.objects.filter(setting=setting).values('parameter__cv_i')[0]['parameter__cv_i']
    cv_a = ParameterUser.objects.filter(setting=setting).values('cv_a')[0]['cv_a']
    rcv = 2 ** 0.5 * (1.96 * (cv_a ** 2 + cv_i ** 2) ** 0.5)
    allow_dev = 0.5 * cv_i
    accept_dev = 0.7 * allow_dev
    allow_bias = 0.25 * math.sqrt(cv_i ** 2 + cv_g ** 2)
    ws3['A23'].font = Font(size=15, bold=True)
    ws3['A24'] = 'Analytical Imprecision (Intra-Assay-CV) (CVa)'
    ws3['B24'] = cv_a
    ws3['A25'] = 'Intra-Individual (Within-subject) Variation (CVi)'
    ws3['B25'] = cv_i
    ws3['A26'] = 'Inter-Individual (Between-subject) Variation (CVg)'
    ws3['B26'] = cv_g
    ws3['A28'] = 'Reference Change Value (RCV%)'
    ws3['B28'] = round(rcv, 2)
    ws3['C28'] = '*1'
    ws3['A29'] = 'Allowable Deviation (%)'
    ws3['B29'] = round(allow_dev, 2)
    ws3['C29'] = '*2'
    ws3['A30'] = 'Acceptable Deviation (%)'
    ws3['B30'] = round(accept_dev, 2)
    ws3['C30'] = '*2'
    ws3['A31'] = 'Allowable Bias (%)'
    ws3['B31'] = round(allow_bias, 2)
    ws3['C31'] = '*2'
    ws3['A33'] = '*1 - Source: doi: 10.1515/cclm-2011-733'
    ws3['A34'] = '*2 - Source: doi: 10.1515/cclm-2019-0596'

    MIN_WIDTH = 10
    for i, column_cells in enumerate(ws3.columns, start=1):
        width = (
            length
            if (length := max(len(str(cell_value) if (cell_value := cell.value) is not None else "")
                              for cell in column_cells)) >= MIN_WIDTH
            else MIN_WIDTH
        )
        ws3.column_dimensions[get_column_letter(i)].width = width

    workbook.save(excelfile)

    response = HttpResponse(excelfile.getvalue(), content_type='application/')
    response['Content-Disposition'] = f'attachment; filename=stability-study-data-{setting.name}-{now}.xlsx'
    return response


# def upload_view(request):
#     form = UploadExcelForm
#     return render(request, 'calculator/upload_form.html', {"form": form})


# --------------------------------------INSTRUMENT-----------------------------------------


def instrument_list(request):
    form = InstrumentForm(request.POST or None)
    instruments = Instrument.objects.all()

    if request.method == 'POST':
        if form.is_valid():

            name = form.cleaned_data["name"]
            manufacturer = form.cleaned_data['manufacturer']
            instrument = Instrument(name=name, manufacturer=manufacturer)
            owner = request.user
            instrument.owner = owner
            instrument.save()
            return redirect('instrument-detail', pk=instrument.id)
        else:
            return render(request, 'calculator/partials/instrument_form.html', context={
                'form': form
            })

    context = {
        'form': form,
        'instruments': instruments,
    }

    return render(request, 'calculator/instrument_list.html', context)


def add_instrument_form(request):
    form = InstrumentForm()
    context = {
        "form": form
    }
    return render(request, 'calculator/partials/instrument_form.html', context)


def instrument_detail(request, pk):
    # instrument = Instrument.objects.get(pk=pk)
    instrument = get_object_or_404(Instrument, pk=pk)
    if not instrument.owner == request.user:
        return HttpResponseForbidden
    else:
        context = {
            "instrument": instrument
        }
        return render(request, 'calculator/partials/instrument_detail.html', context)


def edit_instrument(request, pk):
    instrument = Instrument.objects.get(pk=pk)
    form = InstrumentForm(request.POST or None, instance=instrument)

    # This part is so that the update does not produce more objects
    if request.method == 'POST':
        if form.is_valid():
            instrument = form.save()
            return redirect('instrument-detail', pk=instrument.id)

    context = {
        "form": form,
        "instrument": instrument,
    }
    return render(request, 'calculator/partials/instrument_form.html', context)


def delete_instrument(request, pk):
    instrument = Instrument.objects.get(pk=pk)
    instrument.delete()
    return HttpResponse('')


# ------------------------------------PARAMETER---------------------------------------


def parameter_list(request):
    form = ParameterUserForm(request.POST or None, user=request.user)
    parameters = ParameterUser.objects.all()

    if request.method == 'POST':
        if form.is_valid():
            parameter = form.save(commit=False)
            parameter.owner = request.user
            parameter.save()
            return redirect('parameter-detail', pk=parameter.id)
        else:
            context = {
                'form': form,
                'parameters': ParameterUser.objects.filter()
            }
            return render(request, 'calculator/partials/parameter_form.html', context)

    context = {
        'form': form,
        'parameters': parameters,

    }

    return render(request, 'calculator/parameter_list.html', context)


def search_parameter(request):
    if 'name' in request.POST:
        try:
            parameters = Parameter.objects.filter(name__icontains=request.POST.get("name"))
        except Parameter.DoesNotExist:
            parameters = None

    context = {
        "parameters": parameters,

    }
    return render(request, 'calculator/partials/parameter_searchresult.html', context)


def select_parameter(request, pk):
    parameter = Parameter.objects.get(pk=pk)
    initial_dict = {
        'parameter': parameter.id
    }
    form = ParameterUserForm(
        user=request.user,
        initial=initial_dict
    )

    # form.initial['parameter'].id = parameter.id

    context = {
        "form": form,
        "parameter_name": parameter.name,
        "parameter_unit": parameter.unit,
        "parameter_cvg": parameter.cv_g,
        "parameter_cvi": parameter.cv_i,
    }

    return render(request, 'calculator/partials/parameter_form.html', context)


def add_parameter_form(request):
    # try:
    #     parameter = Parameter.objects.get(name=request.POST.get("name"))
    # except Parameter.DoesNotExist:
    #     parameter = None
    form = ParameterUserForm(user=request.user)
    parameter_select = Parameter.objects.all()
    # form = ParameterUserForm(user=request.user)

    context = {
        "form": form,
        "parameter_list": parameter_select
    }

    return render(request, 'calculator/partials/parameter_form.html', context)


def parameter_detail(request, pk):
    parameter_user = get_object_or_404(ParameterUser, pk=pk)
    if not parameter_user.owner == request.user:
        return HttpResponseForbidden
    else:
        context = {
            "parameter": parameter_user
        }
        return render(request, 'calculator/partials/parameter_detail.html', context)


def edit_parameter(request, pk):
    parameter = ParameterUser.objects.get(pk=pk)
    form = ParameterUserForm(request.POST or None, instance=parameter, user=request.user)

    # This part is so that the update does not produce more objects
    if request.method == 'POST':
        if form.is_valid():
            parameter = form.save()
            return redirect('parameter-detail', pk=parameter.id)

    context = {
        "form": form,
        "parameter": parameter,
        "parameter_name": parameter.parameter.name,
        "parameter_unit": parameter.parameter.unit,
        "parameter_cvg": parameter.parameter.cv_g,
        "parameter_cvi": parameter.parameter.cv_i,
    }
    return render(request, 'calculator/partials/parameter_form.html', context)


def delete_parameter(request, pk):
    parameter = ParameterUser.objects.get(pk=pk)
    parameter.delete()
    return HttpResponse('')


# --------------------------------------PREANALYTICS----------------------------------------

def preanalytical_set_list(request):
    form = PreanalyticalSetForm(request.POST)
    preanalytical_sets = PreanalyticalSet.objects.all()
    if request.method == 'POST':
        if form.is_valid():
            preanalytical_set = form.save(commit=False)
            preanalytical_set.owner = request.user
            preanalytical_set.save()
            return redirect('preanalytical-set-detail', pk=preanalytical_set.id)
        else:
            context = {
                'form': form,
                'preanalytical_sets': PreanalyticalSet.objects.all()
            }
            return render(request, 'calculator/partials/preanalytical_set_form.html', context)

    context = {
        'form': form,
        'preanalytical_sets': preanalytical_sets,

    }

    return render(request, 'calculator/preanalytical_set_list.html', context)


def add_preanalytics_form(request):
    form = PreanalyticalSetForm()
    context = {
        "form": form
    }
    return render(request, 'calculator/partials/preanalytical_set_form.html', context)


def preanalytics_detail(request, pk):
    preanalytical_set = get_object_or_404(PreanalyticalSet, pk=pk)
    if not preanalytical_set.owner == request.user:
        return HttpResponseForbidden
    else:
        context = {
            "preanalytical_set": preanalytical_set
        }
        return render(request, 'calculator/partials/preanalytical_set_detail.html', context)


def edit_preanalytical_set(request, pk):
    preanalytical_set = PreanalyticalSet.objects.get(pk=pk)
    form = PreanalyticalSetForm(request.POST or None, instance=preanalytical_set)

    # This part is so that the update does not produce more objects
    if request.method == 'POST':
        if form.is_valid():
            preanalytical_set = form.save()
            return redirect('preanalytical-set-detail', pk=preanalytical_set.id)

    context = {
        "form": form,
        "preanalytical_set": preanalytical_set,
    }
    return render(request, 'calculator/partials/preanalytical_set_form.html', context)


def delete_preanalytical_set(request, pk):
    preanalytical_set = PreanalyticalSet.objects.get(pk=pk)
    preanalytical_set.delete()
    return HttpResponse('')


# --------------------------------------SAMPLE----------------------------------------

def sample_list(request):
    form = SampleForm(request.POST)
    samples = Sample.objects.all()
    if request.method == 'POST':
        if form.is_valid():
            sample = form.save(commit=False)
            sample.owner = request.user
            sample.save()
            return redirect('sample-detail', pk=sample.id)
        else:
            context = {
                'form': form,
                'samples': Sample.objects.all()
            }
            return render(request, 'calculator/partials/sample_form.html', context)

    context = {
        'form': form,
        'samples': samples,

    }

    return render(request, 'calculator/sample_list.html', context)


def add_sample_form(request):
    form = SampleForm()
    context = {
        "form": form
    }
    return render(request, 'calculator/partials/sample_form.html', context)


def sample_detail(request, pk):
    sample = get_object_or_404(Sample, pk=pk)
    if not sample.owner == request.user:
        return HttpResponseForbidden
    else:
        context = {
            "sample": sample
        }
        return render(request, 'calculator/partials/sample_detail.html', context)


def edit_sample(request, pk):
    sample = Sample.objects.get(pk=pk)
    form = SampleForm(request.POST or None, instance=sample)

    # This part is so that the update does not produce more objects
    if request.method == 'POST':
        if form.is_valid():
            sample = form.save()
            return redirect('sample-detail', pk=sample.id)

    context = {
        "form": form,
        "sample": sample,
    }
    return render(request, 'calculator/partials/sample_form.html', context)


def delete_sample(request, pk):
    sample = Sample.objects.get(pk=pk)
    sample.delete()
    return HttpResponse('')


# -------------------------------------SETTING----------------------------------------

def setting_list(request):
    form = SettingForm(request.POST or None, user=request.user)
    settings = Setting.objects.filter()
    if request.method == 'POST':
        if form.is_valid():
            setting = form.save(commit=False)
            setting.owner = request.user
            setting.save()
            form.save_m2m()
            return redirect('setting-detail', pk=setting.id)
        else:
            context = {
                'form': form,
                'settings': settings
            }
            return render(request, 'calculator/partials/setting_form.html', context)

    context = {
        'form': form,
        'settings': settings

    }

    return render(request, 'calculator/setting_list.html', context)


def add_setting_form(request):
    form = SettingForm(user=request.user)
    owner = request.user

    context = {
        "form": form,
        "owner": owner,

    }
    return render(request, 'calculator/partials/setting_form.html', context)


def setting_detail(request, pk):
    setting = get_object_or_404(Setting, pk=pk)
    if not setting.owner == request.user:
        return HttpResponseForbidden
    else:
        context = {
            "setting": setting,
        }
        return render(request, 'calculator/partials/setting_detail.html', context)


def edit_setting(request, pk):
    setting = Setting.objects.get(pk=pk)
    form = SettingForm(request.POST or None, instance=setting, user=request.user)

    # This part is so that the update does not produce more objects
    if request.method == 'POST':
        if form.is_valid():
            setting = form.save()
            return redirect('setting-detail', pk=setting.id)

    context = {
        "form": form,
        "setting": setting,
    }
    return render(request, 'calculator/partials/setting_form.html', context)


def delete_setting(request, pk):
    setting = Setting.objects.get(pk=pk)
    setting.delete()
    return HttpResponse('')



# -------------------------------------CONDITION----------------------------------------

def condition_list(request):
    form = ConditionForm(request.POST or None)
    conditions = Condition.objects.all()

    if request.method == 'POST':
        if form.is_valid():
            condition = form.save(commit=False)
            condition.owner = request.user
            condition.save()
            return redirect('condition-detail', pk=condition.id)
        else:
            context = {
                'form': form,
                'conditions': Condition.objects.all()
            }
            return render(request, 'calculator/partials/condition_form.html', context)

    context = {
        'form': form,
        'conditions': conditions,

    }

    return render(request, 'calculator/condition_list.html', context)


def add_condition_form(request):
    form = ConditionForm()
    context = {
        "form": form
    }
    return render(request, 'calculator/partials/condition_form.html', context)


def condition_detail(request, pk):
    condition = get_object_or_404(Condition, pk=pk)
    if not condition.owner == request.user:
        return HttpResponseForbidden
    else:
        context = {
            "condition": condition
        }
        return render(request, 'calculator/partials/condition_detail.html', context)


def edit_condition(request, pk):
    condition = Condition.objects.get(pk=pk)
    form = ConditionForm(request.POST or None, instance=condition)

    # This part is so that the update does not produce more objects
    if request.method == 'POST':
        if form.is_valid():
            condition = form.save()
            return redirect('condition-detail', pk=condition.id)

    context = {
        "form": form,
        "condition": condition,
    }
    return render(request, 'calculator/partials/condition_form.html', context)


def delete_condition(request, pk):
    condition = Condition.objects.get(pk=pk)
    condition.delete()
    return HttpResponse('')


# -------------------------------------DURATION----------------------------------------

def duration_list(request):
    form = DurationForm(request.POST or None)
    durations = Duration.objects.all()

    if request.method == 'POST':
        if form.is_valid():
            duration = form.save(commit=False)
            duration.owner = request.user
            duration.save()
            return redirect('duration-detail', pk=duration.id)
        else:
            context = {
                'form': form,
                'durations': Duration.objects.all()
            }
            return render(request, 'calculator/partials/duration_form.html', context)

    context = {
        'form': form,
        'durations': durations,
    }

    return render(request, 'calculator/duration_list.html', context)


def add_duration_form(request):
    form = DurationForm()
    context = {
        "form": form
    }
    return render(request, 'calculator/partials/duration_form.html', context)


def duration_detail(request, pk):
    duration = get_object_or_404(Duration, pk=pk)
    if not duration.owner == request.user:
        return HttpResponseForbidden
    else:
        context = {
            "duration": duration
        }
        return render(request, 'calculator/partials/duration_detail.html', context)


def edit_duration(request, pk):
    duration = Duration.objects.get(pk=pk)
    form = DurationForm(request.POST or None, instance=duration)

    # This part is so that the update does not produce more objects
    if request.method == 'POST':
        if form.is_valid():
            duration = form.save()
            return redirect('duration-detail', pk=duration.id)

    context = {
        "form": form,
        "duration": duration,
    }
    return render(request, 'calculator/partials/duration_form.html', context)


def delete_duration(request, pk):
    duration = Duration.objects.get(pk=pk)
    duration.delete()
    return HttpResponse('')


# -------------------------------------RESULTS---------------------------------

# def create_result(request, setting_pk):
#     setting = Setting.objects.get(pk=setting_pk)
#     durations = Duration.objects.filter(setting=setting)
#     subjects = Subject.objects.filter(setting=setting)
#     results = Result.objects.filter(setting=setting)
#
#     ResultFormSet = modelformset_factory(Result, fields=('value', 'setting', 'duration', 'subject'),
#                                          extra=1)
#
#     if request.method == 'POST':
#         formset = ResultFormSet(request.POST or None)
#
#         if formset.is_valid():
#
#             results = formset.save()
#             # TODO: Show success message after save that fades automatically: https://stackoverflow.com/questions/61153261/make-success-message-disappear-after-few-seconds-of-django-form-submission-and-d
#             # messages.add_message(request, messages.SUCCESS, 'Entry saved')
#             # for result in results:
#             #     sid=result.setting.id
#             # # return redirect('result-detail', pk=rid)
#             # return redirect('create-result', setting_pk=sid)
#
#         else:
#             context = {
#                 'formset': formset,
#                 'results': results
#             }
#             return render(request, 'calculator/partials/result_form.html', context)
#     formset = ResultFormSet(queryset=Result.objects.none())
#
#     context = {
#         'formset': formset,
#         'durations': durations,
#         'setting': setting,
#         'subjects': subjects,
#         'results': results,
#     }
#
#     return render(request, 'calculator/result_list.html', context)
class ResultsFormset(forms.BaseFormSet):
    pass

def result_list(request, setting_pk):
    setting = Setting.objects.get(pk=setting_pk)
    durations = Duration.objects.filter(settings__in=[setting])
    subjects = Subject.objects.filter(settings__in=[setting])
    results = Result.objects.filter(setting=setting)
    # formset = modelformset_factory(
    #     model=Result,
    #     form=ResultForm,
    #     extra=setting.replicate_count*len(subjects),
    #     formset=ResultsFormset,
    # )
    form = ResultForm(subjects, durations, data=request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            # https://zerotobyte.com/using-django-bulk-create-and-bulk-update/
            # https://stackoverflow.com/questions/53594745/what-is-the-use-of-cleaned-data-in-django
            # subject = Result.subject.through(for subject in subjects: return subject.id)
            result_list = []
            for key, value in form.cleaned_data.items():
                if not value:
                    continue
                # print(key, value)
                items = key.split("-")
                subject_id = items[1]
                duration = Duration.objects.get(pk=items[2])
                result = Result(value=value, setting=setting, duration=duration, subject_id=subject_id)
                result.owner = request.user
                result_list.append(result)

            Result.objects.bulk_create(result_list)

            # result = form.save(commit=False)
            # result.setting = setting
            # result.duration = duration
            # # result.subjects.add(subject)
            # # FIXME: Define current subject for asignment
            # result.save()
            # return redirect('result-detail', pk=subject.result.id)
            # return redirect('result-detail')
            form = ResultForm(subjects, durations, data=None)

        else:
            context = {
                'form': form,
                'results': results
            }
            return render(request, 'calculator/partials/result_form.html', context)

    context = {
        "form": form,
        'durations': durations,
        'setting': setting,
        'subjects': subjects,
        'results': results,
    }

    return render(request, 'calculator/result_list.html', context)


def add_result_form(request, setting_pk, duration_pk):
    form = ResultForm()
    # duration = Duration.objects.get(pk=duration_pk)
    subjects = Subject.objects.filter(settings__in=[setting_pk], result__duration_id=duration_pk)

    context = {
        "form": form,
        "subjects": subjects,
        # "duration": duration
    }
    return render(request, 'calculator/partials/result_form.html', context)


def result_detail(request, pk):
    result = get_object_or_404(Result, pk=pk)
    if not result.owner == request.user:
        return HttpResponseForbidden
    else:
        context = {
            "result": result,
            # "duration": duration,
        }
        return render(request, 'calculator/partials/result_detail.html', context)


def edit_result(request, pk):
    result = Result.objects.get(pk=pk)
    form = ResultForm(request.POST or None, instance=result)

    # This part is so that the update does not produce more objects
    if request.method == 'POST':
        if form.is_valid():
            result = form.save()
            return redirect('result-detail', pk=result.id)

    context = {
        "form": form,
        "result": result,
    }
    return render(request, 'calculator/partials/result_form.html', context)


def delete_result(request, pk):
    result = Result.objects.get(pk=pk)
    result.delete()
    return HttpResponse('')


def import_excel(self):
    workbook = load_workbook(result_template_upload)
    print(workbook.sheetnames)
    ws1 = workbook['Basic Info']
    ws2 = workbook['Input results']
    owner_pk = ws1['B18'].value
    setting_pk = ws1["B19"].value
    setting = Setting.objects.get(id=setting_pk)


    # ----get last row and col from input sheet using pandas (openpyxl does not count empty cells)
    resultfile = 'other_data/test.xlsx'
    df = pandas.read_excel(resultfile, sheet_name=[1])
    max_row = 1
    max_col = 1
    for sh_name, sh_content in df.items():
        max_row = len(sh_content) + 1
        max_col = len(sh_content.columns)
    
    def save_results():
        # ----Go through the Excel-File and save results to database/model
        for column in ws2.iter_cols(min_row=5, min_col=4, max_col=max_col - 1, max_row=max_row):
            # print(column)
            for c in column:
                if not setting_pk:
                    break

                c_duration_pk = ws2.cell(row=2, column=c.column).value
                c_subject_pk = ws2.cell(row=c.row, column=2).value
                c_value = c.value

                if not c_value or not c_subject_pk or not c_duration_pk:
                    continue

                # print(c.column)
                # print(c.row)
                # print("owner: " + str(owner_pk))
                # print("setting: " + str(setting_pk))
                # print("value: " + str(c_value))
                # print("subject: " + str(c_subject_pk))
                # print("duration: " + str(c_duration_pk))

                result_object = Result(
                    setting_id=setting_pk,
                    duration_id=c_duration_pk,
                    subject_id=c_subject_pk,
                    value=c_value,
                    owner_id=owner_pk
                )
                result_object.save()

    #---- Check if the setting already has existing results
    if setting.results.exists():
        #---- Delete all existing results for this setting
        for result in setting.results.all():
            result.delete()
        save_results()
    else:
        save_results()

    return HttpResponseRedirect('/calculator/results/' + str(setting_pk))


def result_template_upload(request):
    form = ResultTemplateUploadForm()
    if request.method == 'POST' and request.FILES['result_template_upload_file']:
        uploaded_file = request.FILES['result_template_upload_file']

        workbook = load_workbook(uploaded_file)
        print(workbook.sheetnames)
        ws1 = workbook['Basic Info']
        ws2 = workbook['Input results']
        owner_pk = ws1['B18'].value
        setting_pk = ws1["B19"].value
        setting = Setting.objects.get(id=setting_pk)

        # ----get last row and col from input sheet using pandas (openpyxl does not count empty cells)
        resultfile = 'other_data/test.xlsx'
        df = pandas.read_excel(resultfile, sheet_name=[1])
        max_row = 1
        max_col = 1
        for sh_name, sh_content in df.items():
            max_row = len(sh_content) + 1
            max_col = len(sh_content.columns)

        def save_results():
            # ----Go through the Excel-File and save results to database/model
            for column in ws2.iter_cols(min_row=5, min_col=4, max_col=max_col - 1, max_row=max_row):
                # print(column)
                for c in column:
                    if not setting_pk:
                        break

                    c_duration_pk = ws2.cell(row=2, column=c.column).value
                    c_subject_pk = ws2.cell(row=c.row, column=2).value
                    c_value = c.value

                    if not c_value or not c_subject_pk or not c_duration_pk:
                        continue

                    # print(c.column)
                    # print(c.row)
                    # print("owner: " + str(owner_pk))
                    # print("setting: " + str(setting_pk))
                    # print("value: " + str(c_value))
                    # print("subject: " + str(c_subject_pk))
                    # print("duration: " + str(c_duration_pk))

                    result_object = Result(
                        setting_id=setting_pk,
                        duration_id=c_duration_pk,
                        subject_id=c_subject_pk,
                        value=c_value,
                        owner_id=owner_pk
                    )
                    result_object.save()

        # ---- Check if the setting already has existing results
        if setting.results.exists():
            # ---- Delete all existing results for this setting
            for result in setting.results.all():
                result.delete()
            save_results()
        else:
            save_results()

        # return HttpResponseRedirect('/calculator/results/' + str(setting_pk))
        
        
        
        
        
        

        # import_excel(request.FILES[uploaded_file])
        # fs = FileSystemStorage()
        # filename = fs.save(uploaded_file.name, uploaded_file)
        # uploaded_File_Size = 'Size of Uploaded file: ' + str(uploaded_file.size)
        # content_type_of_uploaded_file = 'Content type of uploaded file: ' + str(uploaded_file.content_type)
        # uploaded_file_name = 'Name of Uploaded file: ' + str(uploaded_file.name)
        messages.success(request, 'Your file "' + str(uploaded_file.name) + '" (' + str(uploaded_file.size) + ' bytes) has been uploaded succesfully')
        message = format_html('<br><a href="results/' + str(setting_pk) + '">Please click here for statistical evaluation of you data</a>')
        messages.success(request, message)

        return render(request, 'calculator/upload_form.html', {"form":form})
    else:
        form = ResultTemplateUploadForm()
    return render(request, 'calculator/upload_form.html', {'form': form})


    #     form = ResultTemplateUploadForm(request.POST, request.FILES)
    #     if form.is_valid():
    #         # form.save()
    #         # df = pandas.read_excel(request.FILES)
    #         # print(df)
    #         import_excel(request.FILES['result_template_upload_file'])
    #         return HttpResponseRedirect('/')
    # else:
    #     form = ResultTemplateUploadForm()
    # return render(request, 'calculator/import_export.html', {'form': form})

# -------------------------------------EXCEL Result Template Download ---------------------------------------


def export_template(request, setting_pk):
    setting = Setting.objects.get(pk=setting_pk)
    if not setting.owner == request.user:
        return HttpResponseForbidden
    else:
        context = {
            "setting": setting
        }
        return render(request, 'calculator/export_template.html', context)

def result_template_download(request, setting_pk):
    exceltemplate = BytesIO()
    workbook = Workbook()
    workbook.remove(workbook.active)
    owner = request.user
    setting = Setting.objects.get(pk=setting_pk, owner=owner)
    subjects = setting.subjects.all()
    durations = setting.durations.all()
    replicates = setting.replicate_count
    parameter_unit = setting.parameter.parameter.unit
    #
    # -----sheet#1 - Basic setting info
    ws1 = workbook.create_sheet(title=f"Basic Info")
    ws1.sheet_view.showGridLines = False

    now = timezone.now().strftime('%d-%m-%Y - %H:%M')

    ws1['A1'] = "EFLM Stability Calculator"
    ws1['A1'].font = Font(size=25, color="e63946", bold=True)
    ws1['A2'] = "EFLM Working Group Preanalytical Phase (WG-PRE)"
    ws1['A2'].font = Font(size=20, color="e63946")
    ws1['A4'] = f"Downloaded by  {request.user.user_name} on {now}"
    ws1['A4'].font = Font(italic=True)
    ws1['A6'] = f"stability study '{setting.name}' Template Sheet"
    ws1['A6'].font = Font(size=20, bold=True, color="1d3557")
    title3 = Font(size=18, color='1d3557')
    for c in ws1['A7:B8']:
        c[0].font = title3
        c[1].font = title3
    ws1['A7'] = "Name"
    ws1['B7'] = str(setting.name)
    ws1['A8'] = "Parameter"
    ws1['B8'] = str(setting.parameter)
    ws1['A9'] = "Samples"
    ws1['B9'] = str(setting.sample)
    ws1['A10'] = "Source of samples"
    ws1['B10'] = str(setting.get_sample_type_display)
    ws1['A11'] = "Storage Condition"
    ws1['B11'] = str(setting.condition)
    ws1['A12'] = "Type of study design"
    ws1['B12'] = str(setting.get_design_type_display())
    ws1['A13'] = "Primary samples or aliquots"
    ws1['B13'] = str(setting.get_design_sample_display())
    ws1['A14'] = "Freeze thaw cycles (n)"
    ws1['B14'] = str(setting.freeze_thaw_cycles)
    ws1['A15'] = "Protocol"
    ws1['B15'] = str(setting.protocol)
    ws1['A16'] = "Comment"
    ws1['B16'] = str(setting.comment)
    ws1["A18"] = "User-ID"
    ws1["B18"] = int(request.user.pk)
    ws1["B18"].alignment = Alignment(horizontal='left')
    ws1["A19"] = "Setting-ID"
    ws1["B19"] = int(setting_pk)
    ws1["B19"].alignment = Alignment(horizontal='left')

    bold = Font(bold=True)
    for c in ws1['A9:A16']:
        c[0].font = bold

    thin = Side(border_style="thin", color="000000")
    bottom_line = Border(bottom=thin)
    for c in ws1['A9:B16']:
        c[0].border = bottom_line
        c[1].border = bottom_line

    MIN_WIDTH = 15
    for i, column_cells in enumerate(ws1.columns, start=1):
        width = (
            length
            if (length := max(len(str(cell_value) if (cell_value := cell.value) is not None else "")
                              for cell in column_cells)) >= MIN_WIDTH
            else MIN_WIDTH
        )
        ws1.column_dimensions[get_column_letter(i)].width = width

    # -------------sheet 2 - data ------------------

    ws2 = workbook.create_sheet(title=f"Input results")
    ws2.sheet_properties.tabColor = "1072BA"
    ws2.sheet_view.showGridLines = True

    ws2["A1"] = 'Storage duration'
    ws2.merge_cells('A1:A2')
    ws2["A1"].alignment = Alignment(vertical="center")
    ws2["B1"] = 'Name'
    ws2["B2"] = 'ID'
    ws2["A3"] = 'Subject'
    ws2.merge_cells('A3:B3')
    ws2["C3"] = 'Replicate'
    ws2.merge_cells('C3:C4')
    ws2["A4"] = 'Name'
    ws2["B4"] = 'ID'

    # ------Add durations in row 1 and seconds in row 2

    i = 0
    for duration in durations:
        ws2.cell(row=1, column =i+4).value = f"{duration.duration_number} {duration.get_duration_unit_display()}"
        # ws2.cell(row=1, column =i+4).font = Font(bold=True)
        # ws2.cell(row=1, column=i + 4).fill = PatternFill(bgColor="1c61ae", fill_type="solid")
        ws2.cell(row=2, column =i+4).value = duration.id
        # ws2.cell(row=2, column =i+4).font = Font(bold=True)
        # ws2.cell(row=2, column=i + 4).fill = PatternFill(bgColor="1c61ae", fill_type="solid")

        i=i+1
    i = 0




    # -----Add subjects and replicates in col 1 and 2
    for subject in subjects:
        for rep in range(0, replicates):
            ws2.cell(row=i + 5, column=1).value = subject.name
            ws2.cell(row=i + 5, column=2).value = subject.id
            ws2.cell(row=i + 5, column=3).value = f"#{rep + 1}"
            i = i + 1
    c=0

    # -----Add parameter unit
    r = 5
    for subject in subjects:
        for replicate in range(0, replicates):
            ws2.cell(row=replicate + r, column=4 + durations.count()).value = parameter_unit
        r = r + replicates
        # c= c + 2

    #-----Format
    white_border = Border(left=Side(style='thin', color="ffffff"),
                          right=Side(style='thin', color="ffffff"),
                          top=Side(style='thin', color="ffffff"),
                          bottom=Side(style='thin', color="ffffff"),
                          )
    black_border = Border(left=Side(style='thin', color="000000"),
                          right=Side(style='thin', color="000000"),
                          top=Side(style='thin', color="000000"),
                          bottom=Side(style='thin', color="000000"),
                          )
    # -format durations
    for col_range in range(1, durations.count() + 4):
        ws2.cell(1, col_range).fill = PatternFill(start_color="1c61ae", end_color="1c61ae", fill_type="solid")
        ws2.cell(2, col_range).fill = PatternFill(start_color="1c61ae", end_color="1c61ae", fill_type="solid")
        ws2.cell(1, col_range).font = Font(bold=True, color="ffffff")
        ws2.cell(2, col_range).font = Font(bold=True, color="ffffff")
        ws2.cell(1, col_range).border = white_border
        ws2.cell(2, col_range).border = white_border

    # -format subjects
    subject_rows = subjects.count()*replicates + 5
    for row_range in range(3, subject_rows):
        ws2.cell(row_range, 1).fill = PatternFill(start_color="62b0df", end_color="62b0df", fill_type="solid")
        ws2.cell(row_range, 2).fill = PatternFill(start_color="62b0df", end_color="62b0df", fill_type="solid")
        ws2.cell(row_range, 1).font = Font(bold=True)
        ws2.cell(row_range, 2).font = Font(bold=True)
        ws2.cell(row_range, 1).border = black_border
        ws2.cell(row_range, 2).border = black_border
        ws2.cell(row_range, 2).alignment = Alignment(horizontal="center")

    # -format replicates
    for row_range in range(3, subject_rows):
        ws2.cell(row_range, 3).fill = PatternFill(start_color="eff6fb", end_color="eff6fb", fill_type="solid")
        ws2.cell(row_range, 3).font = Font(bold=True)
        ws2.cell(row_range, 3).border = black_border
        ws2.cell(row_range, 3).alignment = Alignment(horizontal="center")


  # ---autofit column width
    MIN_WIDTH = 10
    for i, column_cells in enumerate(ws2.columns, start=1):
        width = (
            length
            if (length := max(len(str(cell_value) if (cell_value := cell.value) is not None else "")
                              for cell in column_cells)) >= MIN_WIDTH
            else MIN_WIDTH
        )
        ws2.column_dimensions[get_column_letter(i)].width = width


    workbook.save(exceltemplate)
    response = HttpResponse(exceltemplate.getvalue(), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename=stability-study-{setting.name}-template.xlsx'
    return response



# -------------------------------------EXCEL Result Template UPLOAD ---------------------------------------






# -------------------------------------SUBJECT----------------------------------------

def subject_list(request):
    form = SubjectForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            # subject = form.save(commit=False)
            # subject.owner = request.user
            # subject.save()
            number:int = form.cleaned_data["number"]
            subject_prefix:str = form.cleaned_data["subject_prefix"]

            for n in range(0, number):
                subject: Subject = Subject(
                    name=f"{subject_prefix}_{n+1}",
                    owner=request.user
                )
                subject.save()
            context={'subjects': Subject.objects.filter(owner=request.user)}
            return render(request, 'calculator/subject_list.html', context)
        else:
            context = {
                'form': form,
                'subjects': Subject.objects.filter(owner=request.user)
            }
            return render(request, 'calculator/partials/subject_form.html', context)

    context = {
        'form': form,
        'subjects': Subject.objects.filter(owner=request.user),
    }

    return render(request, 'calculator/subject_list.html', context)


def add_subject_form(request):
    form = SubjectForm()
    context = {
        "form": form,

    }
    return render(request, 'calculator/partials/subject_form.html', context)


def subject_detail(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    if not subject.owner == request.user:
        return HttpResponseForbidden
    else:
        context = {
            "subject": subject
        }
        return render(request, 'calculator/partials/subject_detail.html', context)


def edit_subject(request, pk):
    subject = Subject.objects.get(pk=pk)
    form = SubjectForm(request.POST or None, instance=subject)

    # This part is so that the update does not produce more objects
    if request.method == 'POST':
        if form.is_valid():
            subject = form.save()
            return redirect('subject-detail', pk=subject.id)

    context = {
        "form": form,
        "subject": subject,
    }
    return render(request, 'calculator/partials/subject_form.html', context)


def delete_subject(request, pk):
    subject = Subject.objects.get(pk=pk)
    subject.delete()
    return HttpResponse('')


def item_lists(request):
    settings = Setting.objects.all()
    parameters = Parameter.objects.all()
    subjects = Subject.objects.all()
    durations = Duration.objects.all()
    samples = Sample.objects.all()
    conditions = Condition.objects.all()
    instruments = Instrument.objects.all()

    context = {
        'settings': settings,
        'parameters': parameters,
        'subjects': subjects,
        'durations': durations,
        'samples': samples,
        'conditions': conditions,
        'instruments': instruments,
    }
    return render(request, 'itemlists.html', context)



def new_parameter(request):
    form = NewParameterForm(request.POST or None)
    subject='New Parameter request for Stability Calculator'
    if request.method == 'POST':
        if form.is_valid():
            name = form.cleaned_data['name'],
            unit = form.cleaned_data['unit'],
            cv_a = form.cleaned_data['cv_a'],
            cv_i = form.cleaned_data['cv_i'],
            cv_g = form.cleaned_data['cv_g'],
            user = request.user,
            email = request.user.email,
            context = {
                name:form.cleaned_data['name'],
                unit:form.cleaned_data['unit'],
                cv_a:form.cleaned_data['cv_a'],
                cv_i:form.cleaned_data['cv_i'],
                cv_g:form.cleaned_data['cv_g'],
                user:request.user,
                email:request.user.email,
            }
            # Build the .txt file with the context data just like you build HTML templates!
            plain_message = render_to_string('email_template.txt', context)
            send_mail(subject, plain_message,  'your_account@gmail.com', ['xy@eflm.eu'], fail_silently=False)
            return render(request, 'calculator/thankyou.html')
        # TODO: Add E-Mail Notification
        else:
            context = {
                'form': form,
            }
        return render(request, 'calculator/newparameter.html', context)

        # try:
        #     send_mail(subject, message, from_email, ['admin@example.com'])
        # except BadHeaderError:
        #     return HttpResponse('Invalid header found.')
        # return redirect('success')
    return render(request, "calculator/newparameter.html", {'form': form})


def thankyou_mail(request):
    return HttpResponse('Thank you for your message.')


def SettingAdminList(request):

    filter = SettingFilter(request.GET, queryset=Setting.objects.all())
    return render(request, 'calculator/setting_admin_list.html', {'filter': filter})


def ResultAdminList(request):
    results = Result.objects.all()

    resFilter = ResultFilter(request.GET, queryset=results)
    results = resFilter.qs

    subject_list = []
    for result in results:
        if result.subject in subject_list:
            continue
        else:
            subject_list.append(result.subject)

    subject_count = len(subject_list)

    duration_list = []
    for result in results:
        if result.duration in duration_list:
            continue
        else:
            duration_list.append(result.duration)

    setting_list = []
    for result in results:
        if result.setting in setting_list:

            continue
        else:
            setting_list.append(result.setting)


    # ----------Statistical evaluation------------

    # deviation_dict: dict[int, dict[int, int]] = {}
    # for subject in results.subjects:
    #     if not subject.id in deviation_dict:
    #         deviation_dict[subject.id] = {}
    #         for duration in request.object.durations.all():
    #             if setting == self.object:
    #                 deviation_dict[subject.id][duration.seconds] = subject.deviation(duration, setting)


    # -----Calculate deviation values
    
    
    # ----get values from qs
    results_val = results.values()
    # make dataframe from qs
    results_df= pd.DataFrame(results_val)
    # ---add column headers
    results_df.columns = ['result_id', 'owner_id', 'result', 'setting_id', 'duration_id', 'subject_id', ]
    # ---get seconds values from duration_id
    seconds = []
    for duration in results_df['duration_id']:
        duration = Duration.objects.get(id=duration)
        seconds.append(duration.seconds)
    # ---make an additional column in dataframe with seconds
    results_df["seconds"] = seconds
    # print(results_df)

    # print(results_df.loc[(results_df['seconds'] == 0) & (results_df['setting_id'] == 13)])

    deviations=[]
    for deviation in results_df['result']:
        # setting_id = deviation.setting_id
        baseline = results_df.loc[(results_df['seconds'] == 0) & (results_df['setting_id'] == 13)]
        # result = deviation.value
        # print(baseline['seconds'])
        # print(deviation)
    # results_df = results_df.sort_values(by=['setting_id', 'subject_id'])

    # for row in results_df['setting_id']:
    #     print(results_df[(results_df["seconds"]==0) & (results_df["setting_id"]==13)].result)
    result_groups = results_df.groupby(['setting_id', 'subject_id', 'seconds'])
    average_df = result_groups.aggregate(np.mean).reset_index()
    average_df = average_df.rename(columns={'result':'average'})
    # print(average_df.to_string())
    average_df.columns = ['setting_id', 'subject_id', 'seconds', 'result_id',  'owner_id',  'average',  'duration_id']

    # identify rows with 0
    m = average_df['seconds'].eq(0)
    # compute the sum of rows with 0
    s = (average_df['average'].where(m)
         .groupby([average_df['setting_id'], average_df['subject_id']])
         .sum()
         )

    # compute the deviation per group
    deviation = (
        average_df[['setting_id', 'subject_id']]
            .merge(s, left_on=['setting_id', 'subject_id'], right_index=True, how='left')['average']
            .rdiv(average_df['average']).mul(100)
            .sub(100)
            .round().astype(int)  # optional
            .mask(m, 0)
    )


    average_df['deviation'] = deviation
    # print(average_df)

    # or
    # out = df.assign(deviation=deviation)
    # ---convert x (seconds) into hours
    x1_rel_hours = []
    for x in average_df['seconds']:
        if x == 0:
            x_hour = 0
        else:
            x_hour = x / 3600
        x1_rel_hours.append(x_hour)

    # ---add duration hours to dataframe
    average_df['duration'] = x1_rel_hours

    # print(average_df.to_string())

    y_rel = []
    for y in average_df['deviation']:
        y_rel.append(y)

    # ------preparing data for regression analysis and plotting
    stor_dur = x1_rel_hours
    stor_dur_arr = np.array(stor_dur).reshape(-1, 1)
    stor_dev = y_rel
    stor_dev_arr = np.array(stor_dev).reshape(-1, 1)
    zipped_lin = list(zip(stor_dur, stor_dev))
    df_lin = pd.DataFrame(zipped_lin, columns=['Duration', 'Deviation'])
    # print(stor_dur_arr)
    # print(stor_dev_arr)

    stor_dur_square = []
    for dur in stor_dur:
        dur_square = dur ** 2
        stor_dur_square.append(dur_square)

    zipped_poly = list(zip(stor_dur_square, stor_dev))
    df_poly = pd.DataFrame(zipped_poly, columns=['Duration', 'Deviation'])

    # ---------Linear Regression with SKLearn
    lin_regr = LinearRegression(fit_intercept=False)
    lin_regr.fit(stor_dur_arr, stor_dev_arr)
    prediction = lin_regr.predict(np.sort(stor_dur_arr, axis=0))
    intercept_lin = lin_regr.intercept_
    r2_linregr = lin_regr.score(stor_dur_arr, stor_dev_arr)
    coeff_lin_1 = round(lin_regr.coef_[0][0], 2)


    # -------------Linear regression Graph
    sns.set_style('whitegrid')
    lin_plot = sns.lmplot(x='Duration', y='Deviation', data=df_lin)
    lin_plot_file = BytesIO()
    lin_plot.figure.savefig(lin_plot_file, format='png')
    b64 = base64.b64encode(lin_plot_file.getvalue()).decode()
    # context['chart_lin'] = b64

    # ---------Polynomial Regression 2nd° with SKLearn
    poly = PolynomialFeatures(degree=2,
                              include_bias=False)  # include_bias=False means that we deliberately want the y intercept (ß0) to be equal to 0
    stor_dur_poly = poly.fit_transform(stor_dur_arr)  # create x2 values from our x values
    poly_regr = LinearRegression(fit_intercept=False)
    poly_regr.fit(stor_dur_poly, stor_dev_arr)
    prediction_poly = poly_regr.predict(np.sort(stor_dur_poly, axis=0))
    intercept_poly = poly_regr.intercept_
    r2_polyregr = poly_regr.score(stor_dur_poly, stor_dev_arr)
    coeff_poly_1 = round(poly_regr.coef_[0][0], 2)
    coeff_poly_2 = round(poly_regr.coef_[0][1], 2)


    # -------------Polynomial 2nd° regression Graph
    sns.set_style('whitegrid')
    poly_plot = sns.lmplot(x='Duration', y='Deviation', data=df_lin, order=2)
    poly_plot_file = BytesIO()
    poly_plot.figure.savefig(poly_plot_file, format='png')
    b64_poly = base64.b64encode(poly_plot_file.getvalue()).decode()


    #-----add values to IDs
    
    setting_name = []
    for id in average_df['setting_id']:
        setting = Setting.objects.get(id = id)
        setting_name.append(setting.name)
    average_df['setting_name'] = setting_name

    owner_name = []
    for id in average_df['owner_id']:
        owner = LabUser.objects.get(id=id)
        owner_name.append(owner.user_name)
    average_df['owner_name'] = owner_name

    subject_name = []
    for id in average_df['subject_id']:
        subject = Subject.objects.get(id=id)
        subject_name.append(subject.name)
    average_df['subject_name'] = subject_name


    # -------Print Dataframe to json to render in template
    json_records = average_df.reset_index().to_json(orient='records')
    data = []
    data = json.loads(json_records)
    print(data)

    context = {
        'results' : results,
        'ResultFilter': resFilter,
        'subject_list' : subject_list,
        'duration_list': duration_list,
        'setting_list': setting_list,
        'subject_count': subject_count,
        'r2_linregr' : r2_linregr,
        'eq_linregr' : "PD% = " + str(coeff_lin_1) + " * storage duration",
        'chart_lin' : b64,
        'r2_polyregr' : r2_polyregr,
        'eq_polyregr' : "PD% = " + str(coeff_poly_2) + " * storage duration^2 + " + str(coeff_poly_1) + "* storage duration",
        'chart_poly' : b64_poly,
        'result_table': average_df.to_html(),
        'df': average_df,
        'data': data,
    }

    return render(request, 'calculator/results_admin_list.html', context)


    # def get_context_data(self, **kwargs):
    #     context = super().get_context_data(**kwargs)
    #     context["filter"] = StabilityStudyFilter(self.request.GET, queryset=self.get_queryset())
    #     return context



    
        